import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import time
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
sns.set_theme(style="whitegrid", palette="pastel")



def plot_histograms(df, building_type, gfa_bins, eui_bins, year_built_bins):
    subset_df = df[df['BERDO Property Type'] == building_type]

    fig, axes = plt.subplots(1, 2, figsize=(18, 5))
    fig.suptitle(f'Histograms for {building_type}', fontsize=16)

    # Histogram for Gross Floor Area
    sns.histplot(subset_df['Reported Gross Floor Area (Sq Ft)'], bins=gfa_bins, kde=False, ax=axes[0])
    axes[0].set_title('Reported Gross Floor Area (Sq Ft)')
    axes[0].set_xlabel('Reported Gross Floor Area (Sq Ft)')
    axes[0].set_ylabel('Count')

    # Histogram for Site EUI
    sns.histplot(subset_df['Site EUI (Energy Use Intensity kBtu/ft2)'], bins=eui_bins, kde=False, ax=axes[1])
    axes[1].set_title('Site EUI (Energy Use Intensity kBtu/ft2)')
    axes[1].set_xlabel('Site EUI (Energy Use Intensity kBtu/ft2)')
    axes[1].set_ylabel('Count')

    # # Histogram for Year Built
    # sns.histplot(subset_df['Year Built'], bins=year_built_bins, kde=False, ax=axes[2])
    # axes[2].set_title('Year Built')
    # axes[2].set_xlabel('Year Built')
    # axes[2].set_ylabel('Count')
    #
    # plt.tight_layout(rect=(0, 0, 1, 0.95))
    # plt.show()


def calc_building_type_summary_stats(df, column, bins, labels):
    # Compute summary statistics
    summary_df = df.groupby('BERDO Property Type').agg(
        count=(column, 'count'),
        q1=(column, lambda x: x.quantile(0.25)),
        q2=(column, lambda x: x.quantile(0.50)),  # Median
        q3=(column, lambda x: x.quantile(0.75))
    ).reset_index()

    # Calculate binned counts
    df['bin'] = pd.cut(df[column], bins=bins, labels=labels, right=True)
    binned_counts_df = (df.groupby(['BERDO Property Type', 'bin'], observed=False).size()
                        .unstack(fill_value=0))

    # Ensure that all expected columns are present, even if they are empty
    for label in labels:
        if label not in binned_counts_df.columns:
            binned_counts_df[label] = 0

    binned_counts_df = binned_counts_df.reset_index()

    # Rename columns for clarity
    binned_counts_df.columns = ['BERDO Property Type'] + [f'{label} Count' for label in labels]

    # Join the summary statistics with binned counts
    result_df = pd.merge(summary_df, binned_counts_df, on='BERDO Property Type')

    return result_df


def calc_building_type_allocation(df):
    # Total count and total GFA by building type
    summary_df = df.groupby('BERDO Property Type').agg(
        total_count=('BERDO ID', 'count'),
        total_gfa=('Reported Gross Floor Area (Sq Ft)', 'sum'),
        total_ghg = ('Total GHG Emissions (MT CO2e)', 'sum')
    ).reset_index()

    # Calculate overall totals for all building types
    total_buildings = df['BERDO ID'].count()
    total_floor_area = df['Reported Gross Floor Area (Sq Ft)'].sum()
    total_ghg_emissions = df['Total GHG Emissions (MT CO2e)'].sum()

    # Calculate percentage of total buildings and total GFA
    summary_df['percentage_of_total_buildings'] = (summary_df['total_count'] / total_buildings) * 100
    summary_df['percentage_of_total_gfa'] = (summary_df['total_gfa'] / total_floor_area) * 100
    summary_df['percentage_of_total_ghg'] = (summary_df['total_ghg'] / total_ghg_emissions) * 100

    return summary_df


def filter_and_sort_significant_building_types(df):
    # Calculate total buildings and total GFA
    total_buildings = df['total_count'].sum()
    total_gfa = df['total_gfa'].sum()
    total_ghg = df['total_ghg'].sum()

    # Filter for building types that represent at least 2% of total buildings and GFA
    filtered_df = df[
        (df['total_count'] / total_buildings >= 0.02) &
        (df['total_gfa'] / total_gfa >= 0.02) &
        (df['total_ghg'] / total_ghg >= 0.05)
    ]

    # Sort by total count and then by total GFA (descending order)
    sorted_df = filtered_df.sort_values(by=['total_count', 'total_gfa', 'total_ghg'], ascending=False)

    # Add percentage columns to the DataFrame
    sorted_df['percent_of_total_buildings'] = (sorted_df['total_count'] / total_buildings) * 100
    sorted_df['percent_of_total_gfa'] = (sorted_df['total_gfa'] / total_gfa) * 100
    sorted_df['percent_of_total_ghg'] = (sorted_df['total_ghg'] / total_ghg) * 100

    return sorted_df


def calc_ghg_percentages(df):
    df['percent_of_city_wide_ghg'] = (df['total_ghg'] / TOTAL_CITY_WIDE_EMISSIONS)
    df['percent_of_building_sector_ghg'] = (df['total_ghg'] / TOTAL_BUILDING_SECTOR_EMISSIONS)

    return df


def plot_filtered_building_summary(df, filename):
    fig, axes = plt.subplots(1, 3, figsize=(18, 8))  # figsize=(width, height) in inches

    # Plot for Total Count of Buildings
    sns.barplot(x='total_count', y='BERDO Property Type', data=df, ax=axes[0], palette='Blues_d')
    axes[0].set_title('Total Count of Buildings by Type')
    axes[0].set_xlabel('Total Count')
    axes[0].set_ylabel('Building Type')

    # Extend x-axis to make room for the percentage labels
    max_count = df['total_count'].max()
    axes[0].set_xlim(0, max_count * 1.15)  # Extend the x-axis by 15%

    # Add percentage labels in front of the bars
    for index, value in enumerate(df['total_count']):
        percentage = df['percent_of_total_buildings'].iloc[index]
        axes[0].text(value + (0.02 * max_count), index, f'{percentage:.1f}%', va='center')

    # Plot for Total GFA
    sns.barplot(x='total_gfa', y='BERDO Property Type', data=df, ax=axes[1], palette='Reds_d')
    axes[1].set_title('Total Gross Floor Area (GFA) by Building Type')
    axes[1].set_xlabel('Total GFA (ft²)')
    axes[1].set_ylabel('')

    # Extend x-axis to make room for the percentage labels
    max_gfa = df['total_gfa'].max()
    axes[1].set_xlim(0, max_gfa * 1.15)  # Extend the x-axis by 15%

    # Add percentage labels in front of the bars
    for index, value in enumerate(df['total_gfa']):
        percentage = df['percent_of_total_gfa'].iloc[index]
        axes[1].text(value + (0.02 * max_gfa), index, f'{percentage:.1f}%', va='center')

    # Plot for Total GHG
    sns.barplot(x='total_ghg', y='BERDO Property Type', data=df, ax=axes[2], palette='Greens_d')
    axes[2].set_title('Total GHG Emissions (MT CO2e) by Building Type')
    axes[2].set_xlabel('Total GHG (MT CO2e)')
    axes[2].set_ylabel('')
    tick_interval = 100000
    axes[2].set_xticks(range(0, max(df['total_ghg']) + tick_interval, tick_interval))

    # Extend x-axis to make room for the percentage labels
    max_ghg = df['total_ghg'].max()
    axes[2].set_xlim(0, max_ghg * 1.15)  # Extend the x-axis by 15%

    # Add percentage labels in front of the bars
    for index, value in enumerate(df['total_ghg']):
        percentage = df['percent_of_total_ghg'].iloc[index]
        axes[2].text(value + (0.02 * max_ghg), index, f'{percentage:.1f}%', va='center')

    plt.tight_layout()
    plt.savefig(filename)
    plt.close()


# --------------------------------------------------------------------------------------------------------
# ----------------------------------- City-Wide Emissions Data -------------------------------------------
# --------------------------------------------------------------------------------------------------------

TOTAL_CITY_WIDE_EMISSIONS = 6235970
TOTAL_BUILDING_SECTOR_EMISSIONS = 4335912

# --------------------------------------------------------------------------------------------------------
# ----------------------------------- Clean Data & Generate CSVs -----------------------------------------
# --------------------------------------------------------------------------------------------------------

# File path to BERDO data & read in CSV
file_path_berdo = '../data-files/berdo_data_files/BERDO_Data-Thresholds.csv'
df = pd.read_csv(file_path_berdo)

# file_path to GHG Emissions factors through 2050 and load in CSV
file_path_emissions = '../data-files/berdo_data_files/berdo-emissions-factors.csv'
df_emissions_factors = pd.read_csv(file_path_emissions)

# Subset df to only include relevant columns
columns_to_keep = [
    'BERDO ID', 'Property Owner Name', 'Building Address', 'Reported Gross Floor Area (Sq Ft)', 'Largest Property Type',
    'Site EUI (Energy Use Intensity kBtu/ft2)', 'Total GHG Emissions (MT CO2e)', 'BERDO Property Type',
    'Threshold 2025', 'Threshold 2026', 'Threshold 2027', 'Threshold 2028', 'Threshold 2029', 'Threshold 2030',
    'Threshold 2031', 'Threshold 2032', 'Threshold 2033', 'Threshold 2034', 'Threshold 2035', 'Threshold 2036',
    'Threshold 2037', 'Threshold 2038', 'Threshold 2039', 'Threshold 2040', 'Threshold 2041', 'Threshold 2042',
    'Threshold 2043', 'Threshold 2044', 'Threshold 2045', 'Threshold 2046', 'Threshold 2047', 'Threshold 2048',
    'Threshold 2049', 'Threshold 2050+'
]
df = df[columns_to_keep]
df = df.sort_values(by=['BERDO Property Type'], ascending=True)

# Generate portfolio summary statistics
building_type_summary_df = calc_building_type_allocation(df)
building_type_summary_df = calc_ghg_percentages(building_type_summary_df)
building_type_summary_df.to_csv('../data-files/berdo_data_files/berdo-building_summary.csv', index=False)

# Gross Floor Area (GFA) summary
gfa_bins = [0, 50000, 100000, 250000, 500000, 1000000, float('inf')]
gfa_labels = ['<50k sf', '50k-100k sf', '150k-250k sf', '250k-500k sf', '500k-1M sf', '>1M sf']
gfa_df = calc_building_type_summary_stats(df, 'Reported Gross Floor Area (Sq Ft)', gfa_bins, gfa_labels)
gfa_df.to_csv('../data-files/berdo_data_files/berdo-gfa_summary.csv', index=False)

# Site EUI summary
eui_bins = [0, 20, 40, 60, 80, 100, 150, 250, 500, float('inf')]
eui_labels = ['<20 kbtu/sf', '<40 kbtu/sf', '<60 kbtu/sf', '<80 kbtu/sf', '<100 kbtu/sf', '<150 kbtu/sf',
              '<250 kbtu/sf', '<500 kbtu/sf', '>500 kbtu/sf']
eui_df = calc_building_type_summary_stats(df, 'Site EUI (Energy Use Intensity kBtu/ft2)', eui_bins, eui_labels)
eui_df.to_csv('../data-files/berdo_data_files/berdo-eui_summary.csv', index=False)

# # Year Built summary
# year_built_bins = [0, 1800, 1900, 1940, 1980, 2000, 2010, 2020, float('inf')]
# year_built_labels = ['Built pre-1800', 'Built 1800-1900', 'Built 1900-1940', 'Built 1940-1980', 'Built 1980-2000',
#                      'Built 2000-2010', 'Built 2010-2020', 'Built after 2020']
# year_built_df = calc_building_type_summary_stats(df, 'Year Built', year_built_bins, year_built_labels)
# year_built_df.to_csv('../data-files/berdo_data_files/LL84-year_built_summary.csv', index=False)


# --------------------------------------------------------------------------------------------------------
# ----------------------------------- Excel Summary Stats ------------------------------------------------
# --------------------------------------------------------------------------------------------------------

with pd.ExcelWriter('../data-files/berdo_data_files/berdo_building_summary_statistics.xlsx',
                    engine='openpyxl') as writer:
    building_type_summary_df.to_excel(writer, sheet_name='Building Type Summary', index=False)
    gfa_df.to_excel(writer, sheet_name='GFA Summary', index=False)
    eui_df.to_excel(writer, sheet_name='EUI Summary', index=False)

filtered_sorted_summary_df = filter_and_sort_significant_building_types(building_type_summary_df)

# Generate the plot and save it as an image
plot_filename = '../data-files/berdo_data_files/berdo_building_summary_statistics.png'
plot_filtered_building_summary(filtered_sorted_summary_df, plot_filename)

# Load the workbook and access the active sheet (which will be the first sheet by default)
workbook = load_workbook('../data-files/berdo_data_files/berdo_building_summary_statistics.xlsx')
worksheet = workbook.create_sheet('Summary with Graphs', 0)  # Create a new sheet at the first position

# Insert the image into the sheet
img = Image(plot_filename)
worksheet.add_image(img, 'B2')

# Save the workbook with the image inserted
workbook.save('../data-files/berdo_data_files/berdo_building_summary_statistics.xlsx')

# --------------------------------------------------------------------------------------------------------
# ----------------------------------- Create Histogram Images --------------------------------------------
# --------------------------------------------------------------------------------------------------------

# Grab each different building type
building_types = df['BERDO Property Type'].unique()

# # For loop to create histograms for each building type based on GFA, EUI, and Year Built
# for building_type in building_types:
#     # Create subset df based on building type
#     subset_df = df[df['BERDO Property Type'] == building_type]
#
#     # Create a figure with subplots for GFA, EUI, and Year Built
#     fig, ax = plt.subplots(1, 2, figsize=(18, 5))
#     fig.suptitle(f'Building Type: {building_type}', fontsize=16)
#
#     # Histogram of GFA
#     sns.histplot(subset_df['Reported Gross Floor Area (Sq Ft)'], bins=gfa_bins, kde=False, ax=ax[0])
#     ax[0].set_title('Property GFA (ft2)')
#     ax[0].set_xlabel('Property GFA (ft2)')
#     ax[0].set_ylabel('Count')
#
#     # Histogram of EUI
#     sns.histplot(subset_df['Site EUI (Energy Use Intensity kBtu/ft2)'], bins=eui_bins, kde=False, ax=ax[1])
#     ax[1].set_title('Site EUI (kBtu/sf)')
#     ax[1].set_xlabel('Site EUI (kBtu/sf)')
#     ax[1].set_ylabel('Count')
#
#     # # Histogram of Year Built
#     # sns.histplot(subset_df['Year Built'], bins=year_built_bins, kde=False, ax=ax[2])
#     # ax[2].set_title('Year Built')
#     # ax[2].set_xlabel('Year Built')
#     # ax[2].set_ylabel('Count')
#
#     # Adjust layout to avoid overlapping titles
#     plt.tight_layout(rect=(0.0, 0.0, 1.0, 0.95))
#
#     # Sanitize building type name
#     sanitized_building_type = building_type.replace("/", " or ")
#
#     # Show the plots
#     plt.savefig(f"../images/berdo_summary_stats/plot_{sanitized_building_type}.png")
#     plt.close()  # Close the plot to free up resources
#     time.sleep(0.1)  # Pause for 1 second between plots



